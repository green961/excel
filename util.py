import os
import random

import openpyxl
from faker import Faker

# 创建一个 Faker 对象
fake = Faker()

directory = 'xlsx'


def create_directory(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)
        print(f"Directory '{directory}' created successfully.")
    else:
        pass
        # print(f"Directory '{directory}' already exists.")


create_directory(directory)


def write_excel(i):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '期末成绩'
    titles = ('姓名', '语文', '数学', '英语')
    for col_index, title in enumerate(titles):
        sheet.cell(1, col_index + 1, title)
    names = [fake.name()for _ in range(5)]
    for row_index, name in enumerate(names):
        row = row_index + 2
        sheet.cell(row, 1, name)
        for col_index in range(2, 5):
            sheet.cell(row, col_index, random.randrange(50, 101))

    file_path = os.path.join(directory, f'考试成绩表{i}.xlsx')
    wb.save(file_path)
    return file_path


def get_xlsx_files():
    xlsx_files = []
    for root, _, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            if os.path.isfile(file_path) and os.path.splitext(file_path)[1] == '.xlsx':
                xlsx_files.append(file_path)
    return xlsx_files
